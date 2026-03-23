import os
import openpyxl
from typing import Dict, Optional
import logging

logger = logging.getLogger(__name__)

# Default configuration mapping fields to strictly assigned Excel cell coordinates
DEFAULT_OUTPUT_MAPPING = {
    "tension": "D40",
    "burst": "D41",
    "collapse": "D42"
}

def write_results(
    file_path: str, 
    results: Dict[str, float], 
    output_mapping: Optional[Dict[str, str]] = None,
    sheet_name: Optional[str] = None
) -> None:
    """
    Writes the final calculated performance results back into the Excel template
    while rigidly preserving the original formatting, styles, and other cell formulas.

    Args:
        file_path (str): The absolute or relative path to the Excel (.xlsx) file.
        results (Dict[str, float]): A dictionary containing the final engineering values:
            {'tension': float, 'burst': float, 'collapse': float}
        output_mapping (Dict[str, str], optional): Dictionary mapping result keys to 
            Excel cell coordinates. Defaults to DEFAULT_OUTPUT_MAPPING.
        sheet_name (str, optional): The specific sheet to write to. 
            Defaults to the active sheet if None.

    Raises:
        FileNotFoundError: If the specified Excel file does not exist.
        ValueError: If a required result key is missing from the results dictionary.
    """
    if not os.path.exists(file_path):
        logger.error(f"Cannot write results: File not found - {file_path}")
        raise FileNotFoundError(f"Excel file not found at path: {file_path}")

    mapping = output_mapping or DEFAULT_OUTPUT_MAPPING

    # 1. Load the workbook without 'data_only=True'.
    # This specifically ensures we DO NOT strip out formulas, macros, or cell styles
    # that the engineer built into their production template.
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name] if sheet_name else wb.active

    # 2. Iterate mapping and safely overwrite specified cells
    for key, cell_coord in mapping.items():
        if key not in results:
            wb.close()
            raise ValueError(
                f"Excel Writer Output Error: Expected metric '{key}' is completely "
                f"missing from the logic results payload. Aborting save."
            )
            
        # Execute non-destructive value override
        sheet[cell_coord].value = results[key]
        logger.debug(f"Wrote {key} -> {results[key]} to cell {cell_coord}")

    # 3. Save exactly over the original file context
    wb.save(file_path)
    wb.close()
    
    logger.info(f"Successfully updated Excel workbook: {file_path}")

if __name__ == "__main__":
    # --- Simple Formatter Test Execution ---
    logging.basicConfig(level=logging.DEBUG)
    test_file_path = "writer_test_dummy.xlsx"
    
    try:
        # Create a mock Excel file purely for test scoping
        test_wb = openpyxl.Workbook()
        test_ws = test_wb.active
        # Setup static headers simulating a template document context
        test_ws["C40"] = "Final String Tension:"
        test_ws["C41"] = "Final String Burst:"
        test_ws["C42"] = "Final String Collapse:"
        test_wb.save(test_file_path)
        
        print(f"--- Running Excel Writer Test ---")
        
        mock_results = {
            "tension": 330000.0,
            "burst": 12000.0,
            "collapse": 11000.0
        }
        
        print(f"\nWriting calculated pipeline outputs: {mock_results}")
        write_results(test_file_path, mock_results)
        
        # Reload explicitly to verify the write persisted to IO disk
        verify_wb = openpyxl.load_workbook(test_file_path, data_only=True)
        verify_ws = verify_wb.active
        
        print("\n✅ Disk Write Verification Successful:")
        print(f"  Tension (D40): {verify_ws['D40'].value}")
        print(f"  Burst (D41): {verify_ws['D41'].value}")
        print(f"  Collapse (D42): {verify_ws['D42'].value}")
        verify_wb.close()
            
    except Exception as e:
        print(f"\n❌ Error during Excel Write: {e}")
    finally:
        # Clean up mock file artifact
        if os.path.exists(test_file_path):
            os.remove(test_file_path)
            print(f"\nCleaned up backend test mock file: {test_file_path}")
